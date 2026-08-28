# -*- coding: utf-8 -*-
"""離線自我驗證：不需要網路、不會開瀏覽器。

用假造的原始資料檔跑一遍「過濾 → 換算 → 產出 Excel → 校正參數」，
確認打包前修掉的問題不會再回來。

用法：
    python tools/selftest.py
"""

import os
import shutil
import sys
import tempfile
import zipfile
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import tiltdotdatatransfer as app  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

PASSED = []
FAILED = []


def check(name, cond, detail=""):
    if cond:
        PASSED.append(name)
        print(f"  [PASS] {name}")
    else:
        FAILED.append(f"{name} {detail}".strip())
        print(f"  [FAIL] {name} {detail}")


# ---------- 假資料 ----------
COL_X, COL_Y, COL_Z, COL_T = 4, 5, 6, 11  # 0-based


def make_rows(n=24, date="2026/08/12", bad_row=None, blank_first=False):
    """溫度先升後降，X/Y/Z 隨溫度線性漂移。"""
    rows = []
    for i in range(n):
        t = 20.0 + (i if i < n // 2 else n - i) * 1.0
        x = 0.010 + (t - 20.0) * 0.0002
        y = -0.005 + (t - 20.0) * 0.0001
        z = 0.002 + (t - 20.0) * 0.00005
        hh, mm = divmod(i * 30, 60)
        stamp_time = f"{7 + hh:02d}:{mm:02d}:00"
        if bad_row is not None and i == bad_row:
            x = 99.0  # 換算後 > 10000 arcsec
        if blank_first and i == 0:
            xs, ys, zs, ts = "", "", "", ""
        else:
            xs, ys, zs, ts = f"{x:.6f}", f"{y:.6f}", f"{z:.6f}", f"{t:.2f}"
        rows.append(["36901", date, stamp_time, "OK", xs, ys, zs, "3.7", "-90", "0", "0", ts])
    return rows


def write_csv(path, rows, delim=",", header=True):
    lines = []
    if header:
        lines.append(delim.join(["SN", "Date", "Time", "Status", "X", "Y", "Z",
                                 "Batt", "RSSI", "R1", "R2", "Temp"]))
    for r in rows:
        lines.append(delim.join(r))
    with open(path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    return path


def dt(s):
    return datetime.strptime(s, "%Y-%m-%d %H:%M:%S")


def media_count(xlsx_path):
    with zipfile.ZipFile(xlsx_path) as z:
        return len([n for n in z.namelist() if n.startswith("xl/media/")])


# ---------- 測試 ----------
def test_input_parsing():
    print("\n[1] 輸入解析")
    check("parse_serials 範圍展開", app.parse_serials("36901-36903") == ["36901", "36902", "36903"])
    check("parse_serials 混合輸入", app.parse_serials("36901-36902, 36910") == ["36901", "36902", "36910"])
    check("parse_serials 起訖顛倒自動對調", app.parse_serials("36903-36901") == ["36901", "36902", "36903"])
    check("parse_serials 去重", app.parse_serials("36901,36901") == ["36901"])
    check("normalize_date 補零", app.normalize_date("2026/8/2") == "2026-08-02")
    check("normalize_time 補秒", app.normalize_time("7:45") == "07:45:00")
    check("normalize_time 補零", app.normalize_time("7:5:3") == "07:05:03")
    check("無效日期原樣回傳", app.normalize_date("2026-13-45") == "2026-13-45")
    check("_valid_date 擋掉無效日期", not app._valid_date("2026-13-45"))


def test_date_filter(tmp):
    print("\n[2] 時間範圍過濾（B7 / B8）")
    padded = write_csv(os.path.join(tmp, "padded.csv"), make_rows(date="2026/08/12"))
    unpadded = write_csv(os.path.join(tmp, "unpadded.csv"), make_rows(date="2026/8/12"))

    s, e = dt("2026-08-12 08:00:00"), dt("2026-08-12 12:00:00")
    t1, _, i1 = app._filter_data_lines(padded, s, e)
    t2, _, i2 = app._filter_data_lines(unpadded, s, e)
    check("未補零日期與補零日期結果一致", i1["kept"] == i2["kept"] and i1["kept"] > 0,
          f"(padded={i1['kept']}, unpadded={i2['kept']})")
    check("過濾後不含表頭", "Date" not in t1.splitlines()[0])
    check("過濾筆數少於總筆數", 0 < i1["kept"] < i1["total"], f"({i1['kept']}/{i1['total']})")

    # 使用者輸入 7:45 與 07:45:00 應等價
    a = app.normalize_time("7:45")
    _, _, ia = app._filter_data_lines(padded, dt(f"2026-08-12 {a}"), e)
    _, _, ib = app._filter_data_lines(padded, dt("2026-08-12 07:45:00"), e)
    check("7:45 與 07:45:00 等價", ia["kept"] == ib["kept"])

    # DMY / MDY
    dmy = write_csv(os.path.join(tmp, "dmy.csv"), make_rows(date="12/08/2026"))
    app.CONFIG["date_order"] = "auto"
    _, _, i3 = app._filter_data_lines(dmy, s, e)
    check("DD/MM/YYYY 可解析", i3["kept"] == i1["kept"], f"({i3['kept']} vs {i1['kept']})")

    app.CONFIG["date_order"] = "MDY"
    mdy = write_csv(os.path.join(tmp, "mdy.csv"), make_rows(date="08/12/2026"))
    _, _, i4 = app._filter_data_lines(mdy, s, e)
    check("date_order=MDY 設定生效", i4["kept"] == i1["kept"], f"({i4['kept']} vs {i1['kept']})")
    app.CONFIG["date_order"] = "auto"

    # 完全在範圍外
    _, _, i5 = app._filter_data_lines(padded, dt("2020-01-01 00:00:00"), dt("2020-01-02 00:00:00"))
    check("範圍外回傳 0 筆", i5["kept"] == 0)


def test_delimiter(tmp):
    print("\n[3] 分隔符號（B11）")
    semi = write_csv(os.path.join(tmp, "semi.csv"), make_rows(), delim=";")
    s, e = dt("2026-08-12 08:00:00"), dt("2026-08-12 09:00:00")
    text, delim, info = app._filter_data_lines(semi, s, e)
    check("分號檔偵測為 ;", delim == ";", f"(得到 {delim!r})")
    check("過濾後僅剩少數列", info["kept"] <= 3, f"({info['kept']})")
    out, _, _ = app.transform_text_to_xlsx(text, delim, os.path.join(tmp, "36901_semi.txt"),
                                           "XYZ", COL_X, COL_Y, COL_Z, COL_T)
    wb = load_workbook(out)
    ws = wb.active
    check("短資料仍用原分隔符解析（欄位未錯位）", ws["E1"].value == "X" and ws["L1"].value == "T",
          f"(E1={ws['E1'].value}, L1={ws['L1'].value})")

    pipe = write_csv(os.path.join(tmp, "pipe.csv"), make_rows(), delim="|")
    _, d2, _ = app._filter_data_lines(pipe, s, e)
    check("直線分隔可偵測", d2 == "|", f"(得到 {d2!r})")


def test_transform_transfer(tmp):
    print("\n[4] TRANSFER 模式輸出")
    src = write_csv(os.path.join(tmp, "36901_raw.txt"), make_rows())
    s, e = dt("2026-08-12 00:00:00"), dt("2026-08-13 00:00:00")
    text, delim, _ = app._filter_data_lines(src, s, e)
    out, outliers, base_idx = app.transform_text_to_xlsx(text, delim, src, "XYZ",
                                                         COL_X, COL_Y, COL_Z, COL_T)
    wb = load_workbook(out)
    ws = wb.active
    check("標題列標記 X/Y/Z/T", [ws["E1"].value, ws["F1"].value, ws["G1"].value, ws["L1"].value]
          == ["X", "Y", "Z", "T"])
    check("換算欄標題", [ws["O1"].value, ws["P1"].value, ws["Q1"].value, ws["R1"].value]
          == ["X(sec)", "Y(sec)", "Z(sec)", "ΔT"])
    check("無補償欄", ws["S1"].value in (None, ""))
    check("換算公式正確", ws["O2"].value == "=(E2-$E$2)*3600", f"(得到 {ws['O2'].value})")
    check("ΔT 公式正確", ws["R2"].value == "=L2-$L$2", f"(得到 {ws['R2'].value})")
    check("基準列為第 0 列", base_idx == 0)
    check("無異常紀錄", outliers == [])
    check("嵌入 1 張圖", media_count(out) == 1, f"({media_count(out)})")
    check("圖表不覆蓋資料（錨點在資料右側）",
          all(img.anchor._from.col >= 18 for img in ws._images) if ws._images else True)
    check("檔名未加異常前綴", not os.path.basename(out).startswith("[異常]"))
    return out


def test_transform_test_mode(tmp):
    print("\n[5] TEST 模式輸出（含補償欄）")
    src = write_csv(os.path.join(tmp, "36902_raw.txt"), make_rows())
    s, e = dt("2026-08-12 00:00:00"), dt("2026-08-13 00:00:00")
    text, delim, _ = app._filter_data_lines(src, s, e)
    out, _, _ = app.transform_text_to_xlsx(text, delim, src, "XYZ",
                                           COL_X, COL_Y, COL_Z, COL_T, run_mode="TEST")
    wb = load_workbook(out)
    ws = wb.active
    check("補償欄標題", [ws["S1"].value, ws["T1"].value, ws["U1"].value]
          == ["補償後X(sec)", "補償後Y(sec)", "補償後Z(sec)"])
    formula = ws["S2"].value
    check("補償公式格式正確", formula.startswith("=(E2-$E$2+(") and formula.endswith("*R2))*3600"),
          f"(得到 {formula})")
    check("補償係數不是 nan", "nan" not in formula, f"(得到 {formula})")
    check("嵌入 2 張圖", media_count(out) == 2, f"({media_count(out)})")

    # XY 模式
    out_xy, _, _ = app.transform_text_to_xlsx(text, delim, os.path.join(tmp, "36903_raw.txt"),
                                              "XY", COL_X, COL_Y, None, COL_T, run_mode="TEST")
    ws_xy = load_workbook(out_xy).active
    check("XY 模式沒有 Z 欄", ws_xy["Q1"].value == "ΔT" and ws_xy["R1"].value == "補償後X(sec)",
          f"(Q1={ws_xy['Q1'].value}, R1={ws_xy['R1'].value})")
    return out


def test_outliers_and_base(tmp):
    print("\n[6] 異常值與基準列（B13 / B14）")
    src = write_csv(os.path.join(tmp, "36904_raw.txt"), make_rows(bad_row=5))
    s, e = dt("2026-08-12 00:00:00"), dt("2026-08-13 00:00:00")
    text, delim, _ = app._filter_data_lines(src, s, e)
    out, outliers, _ = app.transform_text_to_xlsx(text, delim, src, "XYZ",
                                                  COL_X, COL_Y, COL_Z, COL_T)
    check("偵測到異常值", len(outliers) == 1, f"({outliers})")
    check("異常紀錄含日期時間", "2026/08/12" in outliers[0] and ":" in outliers[0], f"({outliers[0]})")
    check("異常紀錄指向 Excel 第 7 列", "第 7 列" in outliers[0], f"({outliers[0]})")
    check("檔名加上異常前綴", os.path.basename(out).startswith("[異常]"))

    ws = load_workbook(out).active
    check("異常列原始值被清空", (ws["E7"].value in (None, "")) and (ws["L7"].value in (None, "")),
          f"(E7={ws['E7'].value})")
    check("異常列換算欄被清空", ws["O7"].value in (None, ""))
    check("非異常列仍有公式", ws["O8"].value == "=(E8-$E$2)*3600", f"(得到 {ws['O8'].value})")

    # 第一列無數值時，圖表與公式要用同一個基準列
    src2 = write_csv(os.path.join(tmp, "36905_raw.txt"), make_rows(blank_first=True))
    text2, delim2, _ = app._filter_data_lines(src2, s, e)
    out2, _, base_idx2 = app.transform_text_to_xlsx(text2, delim2, src2, "XYZ",
                                                    COL_X, COL_Y, COL_Z, COL_T)
    ws2 = load_workbook(out2).active
    check("第一列空白時基準列往後移", base_idx2 == 1, f"(base_idx={base_idx2})")
    check("Excel 公式基準列與 base_idx 一致", ws2["O3"].value == "=(E3-$E$3)*3600",
          f"(得到 {ws2['O3'].value})")
    return out


def test_compensation_effect(tmp):
    print("\n[11] 補償係數確實能抵銷溫漂")
    import pandas as pd
    src = write_csv(os.path.join(tmp, "36908_raw.txt"), make_rows())
    s, e = dt("2026-08-12 00:00:00"), dt("2026-08-13 00:00:00")
    text, delim, _ = app._filter_data_lines(src, s, e)
    out, _, base_idx = app.transform_text_to_xlsx(text, delim, src, "XYZ",
                                                  COL_X, COL_Y, COL_Z, COL_T)
    df = pd.read_excel(out)
    T = app._numeric(df, COL_T)
    X = app._numeric(df, COL_X)
    mask = T.notna() & X.notna()
    b = mask.idxmax()
    c = app._c_from_peak(X, float(X.loc[b]), T, float(T.loc[b]), mask)

    raw_peak = ((X - float(X.loc[b])) * 3600).abs().max()
    comp = (X - float(X.loc[b]) + c * (T - float(T.loc[b]))) * 3600
    comp_peak = comp.abs().max()
    check("未補償時有明顯溫漂", raw_peak > 5, f"(peak={raw_peak:.1f} arcsec)")
    check("補償後峰值大幅下降", comp_peak < raw_peak * 0.05,
          f"(raw={raw_peak:.1f} → comp={comp_peak:.4f})")
    check("補償係數符號正確", c < 0, f"(c={c})")


def test_calibration(tmp, files):
    print("\n[7] 校正參數總表（B16 / B17）")
    out_dir = app.PROCESSED_DIR
    path = app.generate_calibration_parameters(files, out_dir, "XYZ", COL_X, COL_Y, COL_Z, COL_T)
    check("總表產生成功", path is not None and os.path.exists(path))
    if path:
        import pandas as pd
        df = pd.read_excel(path)
        check("總表列數等於傳入檔案數", len(df) == len(files), f"({len(df)} vs {len(files)})")
        check("含補償係數欄位",
              {"基準溫度(T_base)", "X軸補償係數 (cx)", "Y軸補償係數 (cy)", "Z軸補償係數 (cz)"}
              <= set(df.columns), f"({list(df.columns)})")
        check("係數為有限值", df["X軸補償係數 (cx)"].notna().all())
        stale = os.path.join(out_dir, "舊的_不該被統計.xlsx")
        shutil.copy(files[0], stale)
        path2 = app.generate_calibration_parameters(files, out_dir, "XYZ",
                                                    COL_X, COL_Y, COL_Z, COL_T)
        df2 = pd.read_excel(path2)
        check("舊檔不會被混入總表", len(df2) == len(files), f"({len(df2)} vs {len(files)})")


def test_report_and_guards(tmp):
    print("\n[8] 報告與防呆（B6 / B12 / B15）")
    app.PROCESSED_DIR = os.path.join(tmp, "不存在的資料夾", "修改過的")
    report = app.write_report({}, [("36901", "下載逾時或網站無回應")], [])
    check("資料夾不存在時仍能寫出報告", os.path.exists(report))
    content = open(report, encoding="utf-8-sig").read()
    check("報告含無資料序號", "36901" in content)
    check("報告分開列出轉換失敗", "【已下載但轉換失敗的序號】" not in content)

    report2 = app.write_report({"36902": ["X軸 ... 異常"]},
                               [], [("36903", "欄位設定超出範圍")])
    c2 = open(report2, encoding="utf-8-sig").read()
    check("報告含轉換失敗區塊", "【已下載但轉換失敗的序號】" in c2 and "36903" in c2)
    check("報告含異常紀錄區塊", "36902" in c2)

    check("nan 係數被歸零", app._safe_coef(float("nan"), "X軸") == 0.0)
    check("inf 係數被歸零", app._safe_coef(float("inf"), "Y軸") == 0.0)
    check("正常係數保留", abs(app._safe_coef(-0.00012345, "Z軸") + 0.00012345) < 1e-12)


def test_bad_column_setting(tmp):
    print("\n[9] 欄位設定錯誤的提示")
    src = write_csv(os.path.join(tmp, "36906_raw.txt"), make_rows())
    s, e = dt("2026-08-12 00:00:00"), dt("2026-08-13 00:00:00")
    text, delim, _ = app._filter_data_lines(src, s, e)
    try:
        app.transform_text_to_xlsx(text, delim, src, "XYZ", COL_X, COL_Y, COL_Z, 99)
        check("欄位超界應丟出明確錯誤", False, "(沒有丟出例外)")
    except ValueError as e:
        check("欄位超界丟出明確錯誤", "欄位設定超出原始資料範圍" in str(e), f"({e})")

    # 指定到非數值欄位
    try:
        app.transform_text_to_xlsx(text, delim, src, "XYZ", 3, 3, 3, 3)
        check("全非數值欄位應丟出明確錯誤", False, "(沒有丟出例外)")
    except ValueError as e:
        check("全非數值欄位丟出明確錯誤", "找不到任何數值資料" in str(e), f"({e})")


def test_formula_text_guard(tmp):
    print("\n[10] 以 = 開頭的儲存格不會被當成公式")
    rows = make_rows(n=4)
    rows[1][3] = "=SUM(1)"
    src = write_csv(os.path.join(tmp, "36907_raw.txt"), rows)
    s, e = dt("2026-08-12 00:00:00"), dt("2026-08-13 00:00:00")
    text, delim, _ = app._filter_data_lines(src, s, e)
    out, _, _ = app.transform_text_to_xlsx(text, delim, src, "XYZ", COL_X, COL_Y, COL_Z, COL_T)
    ws = load_workbook(out).active
    check("文字型 = 開頭儲存格保持文字", ws["D3"].data_type == "s", f"(data_type={ws['D3'].data_type})")


def test_config_and_dirs(tmp):
    print("\n[12] 設定檔與輸出資料夾（A1 / A2 / A5）")
    import json
    cfg_dir = os.path.join(tmp, "app")
    os.makedirs(cfg_dir, exist_ok=True)
    original_app_dir = app.APP_DIR
    try:
        app.APP_DIR = cfg_dir
        cfg = app.load_config()
        cfg_file = os.path.join(cfg_dir, app.CONFIG_FILENAME)
        check("首次執行自動產生設定檔", os.path.exists(cfg_file))
        check("預設值正確", cfg["outlier_threshold"] == 10000.0 and cfg["date_col"] == 2)

        # 使用者手改設定檔
        data = json.load(open(cfg_file, encoding="utf-8"))
        data["outlier_threshold"] = 500
        data["date_order"] = "MDY"
        data["download_timeout_ms"] = 100          # 太小，應被拉回下限
        data["date_col"] = 0                       # 不合法，應退回預設
        json.dump(data, open(cfg_file, "w", encoding="utf-8"), ensure_ascii=False)
        cfg = app.load_config()
        check("設定檔可覆寫閾值", cfg["outlier_threshold"] == 500.0)
        check("設定檔可指定日期順序", cfg["date_order"] == "MDY")
        check("不合法逾時被拉回下限", cfg["download_timeout_ms"] == 5000)
        check("不合法欄位編號退回預設", cfg["date_col"] == 2)

        # 壞掉的 JSON 不應該讓程式中斷
        open(cfg_file, "w", encoding="utf-8").write("{ 這不是 JSON")
        cfg = app.load_config()
        check("設定檔損壞時沿用預設值", cfg["outlier_threshold"] == 10000.0)

        # 輸出資料夾
        app.CONFIG["output_root"] = ""
        app.setup_output_dirs(interactive=False)
        check("輸出預設建在 exe 所在資料夾",
              app.PROCESSED_DIR == os.path.join(cfg_dir, "修改過的"), f"({app.PROCESSED_DIR})")

        custom = os.path.join(tmp, "自訂輸出")
        app.CONFIG["output_root"] = custom
        app.setup_output_dirs(interactive=False)
        check("設定檔可指定輸出位置",
              app.PROCESSED_DIR == os.path.join(custom, "修改過的"), f"({app.PROCESSED_DIR})")
        check("_ensure_writable 對可寫路徑回 True", app._ensure_writable(custom))
        check("_ensure_writable 對不可寫路徑回 False",
              not app._ensure_writable("/proc/不可能建立的資料夾/x"))
    finally:
        app.APP_DIR = original_app_dir
        app.CONFIG = dict(app.DEFAULT_CONFIG)


def main():
    tmp = tempfile.mkdtemp(prefix="tiltdot_selftest_")
    app.CONFIG = dict(app.DEFAULT_CONFIG)
    app.PROCESSED_DIR = os.path.join(tmp, "修改過的")
    app.RAW_DIR = os.path.join(tmp, "初始檔案")
    print(f"暫存資料夾：{tmp}")

    try:
        test_input_parsing()
        test_date_filter(tmp)
        test_delimiter(tmp)
        f1 = test_transform_transfer(tmp)
        f2 = test_transform_test_mode(tmp)
        f3 = test_outliers_and_base(tmp)
        test_calibration(tmp, [f1, f2, f3])
        test_compensation_effect(tmp)
        test_bad_column_setting(tmp)
        test_formula_text_guard(tmp)
        test_report_and_guards(tmp)
        test_config_and_dirs(tmp)
    finally:
        print("\n" + "=" * 50)
        print(f"通過 {len(PASSED)} 項，失敗 {len(FAILED)} 項")
        for f in FAILED:
            print(f"  - {f}")
        print("=" * 50)
        shutil.rmtree(tmp, ignore_errors=True)

    return 1 if FAILED else 0


if __name__ == "__main__":
    sys.exit(main())
