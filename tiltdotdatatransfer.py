# -*- coding: utf-8 -*-
"""
tiltDot 資料下載與轉換工具

從 WMS 網站下載指定序號的傾斜計原始資料，依時間範圍過濾後換算成角秒 (arcsec)，
產出含圖表的 Excel 檔，並統整各感測器的溫度補償係數。

執行模式：
  TRANSFER  只做換算、異常過濾、產出未補償圖表與校正係數總表
  TEST      在 TRANSFER 的基礎上，額外產出補償後欄位與補償後圖表
"""

import asyncio
import json
import math
import os
import re
import sys
import traceback
from datetime import datetime
from io import BytesIO, StringIO

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import column_index_from_string, get_column_letter
from PIL import Image as PILImage

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

# Playwright 只有「下載」流程需要；離線處理與自我測試不應該因為缺它而無法匯入。
try:
    from playwright.async_api import async_playwright
except Exception:  # pragma: no cover - 取決於執行環境
    async_playwright = None


# ---------- 主控台編碼 ----------
def _reconfigure_stdio():
    """避免輸出被導向檔案時，中文與 emoji 觸發 cp950 的 UnicodeEncodeError。"""
    for stream in (sys.stdout, sys.stderr):
        try:
            stream.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass


# ---------- 路徑 ----------
def _app_dir() -> str:
    """回傳 exe（打包後）或 .py（開發中）所在的資料夾。

    PyInstaller 下 __file__ 會指向解壓後的暫存目錄，不能拿來當輸出位置。
    """
    if getattr(sys, "frozen", False):
        return os.path.dirname(os.path.abspath(sys.executable))
    return os.path.dirname(os.path.abspath(__file__))


APP_DIR = _app_dir()
CONFIG_FILENAME = "tiltdot_config.json"

# 執行期才決定的輸出位置（由 setup_output_dirs() 設定）
RAW_DIR = os.path.join(APP_DIR, "初始檔案")
PROCESSED_DIR = os.path.join(APP_DIR, "修改過的")


# ---------- 設定檔 ----------
DEFAULT_CONFIG = {
    "_說明": "date_col / time_col 為原始檔中日期與時間的欄位編號，1 代表第一欄(A)。修改後存檔即可生效。",
    "url": "https://wms.sanlien.com.tw/dotadmin/view/view_mqtt.html",
    "output_root": "",
    "remove_outliers": True,
    "outlier_threshold": 10000.0,
    "headless": False,
    "default_wait_ms": 4000,
    "download_timeout_ms": 30000,
    "date_col": 2,
    "time_col": 3,
    "date_order": "auto",
}

CONFIG = dict(DEFAULT_CONFIG)


def _ensure_writable(path: str) -> bool:
    """確認資料夾存在且可寫入。"""
    try:
        os.makedirs(path, exist_ok=True)
        probe = os.path.join(path, ".write_test")
        with open(probe, "w", encoding="utf-8") as f:
            f.write("ok")
        os.remove(probe)
        return True
    except OSError:
        return False


def _fallback_config_dir() -> str:
    base = os.environ.get("LOCALAPPDATA") or os.path.join(os.path.expanduser("~"), ".config")
    return os.path.join(base, "tiltdot")


def _config_path() -> str:
    """設定檔位置：優先放 exe 旁邊，該處不可寫時退到使用者設定目錄。"""
    if _ensure_writable(APP_DIR):
        return os.path.join(APP_DIR, CONFIG_FILENAME)
    return os.path.join(_fallback_config_dir(), CONFIG_FILENAME)


def _coerce(value, default):
    try:
        if isinstance(default, bool):
            if isinstance(value, str):
                return value.strip().lower() in ("1", "true", "yes", "y", "on")
            return bool(value)
        if isinstance(default, float):
            return float(value)
        if isinstance(default, int):
            return int(value)
        if isinstance(default, str):
            return str(value)
    except (TypeError, ValueError):
        return default
    return value


def load_config() -> dict:
    """讀取設定檔；不存在就建立一份預設值，內容壞掉則沿用預設值並警告。"""
    global CONFIG
    cfg = dict(DEFAULT_CONFIG)
    path = _config_path()
    if os.path.exists(path):
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
            if isinstance(data, dict):
                for key, default in DEFAULT_CONFIG.items():
                    if key in data:
                        cfg[key] = _coerce(data[key], default)
            else:
                print(f"[警告] 設定檔格式不正確，改用預設值：{path}")
        except Exception as e:
            print(f"[警告] 設定檔讀取失敗（{e}），改用預設值：{path}")
    else:
        CONFIG = cfg
        save_config(cfg)

    cfg["_說明"] = DEFAULT_CONFIG["_說明"]
    if cfg["date_col"] < 1:
        cfg["date_col"] = DEFAULT_CONFIG["date_col"]
    if cfg["time_col"] < 1:
        cfg["time_col"] = DEFAULT_CONFIG["time_col"]
    if cfg["date_order"].upper() not in ("AUTO", "DMY", "MDY"):
        cfg["date_order"] = "auto"
    if cfg["download_timeout_ms"] < 5000:
        cfg["download_timeout_ms"] = 5000
    if cfg["outlier_threshold"] <= 0:
        cfg["outlier_threshold"] = DEFAULT_CONFIG["outlier_threshold"]

    CONFIG = cfg
    return cfg


def save_config(cfg: dict) -> None:
    path = _config_path()
    try:
        os.makedirs(os.path.dirname(path), exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            json.dump(cfg, f, ensure_ascii=False, indent=2)
    except OSError as e:
        print(f"[警告] 設定檔寫入失敗（{e}），本次修改不會被記住。")


def setup_output_dirs(interactive: bool = True) -> None:
    """決定「初始檔案」「修改過的」的位置，必要時退到使用者文件夾。"""
    global RAW_DIR, PROCESSED_DIR

    root = (CONFIG.get("output_root") or "").strip() or APP_DIR
    if not _ensure_writable(root):
        fallback = os.path.join(os.path.expanduser("~"), "Documents", "tiltdot")
        print(f"[提醒] 原輸出位置沒有寫入權限：{root}")
        print(f"        已自動改為：{fallback}")
        root = fallback
        _ensure_writable(root)

    if interactive:
        print(f"\n目前輸出資料夾：{root}")
        try:
            answer = input("要改到其他資料夾嗎？輸入完整路徑，直接按 Enter 沿用：").strip().strip('"')
        except EOFError:
            answer = ""
        if answer:
            if _ensure_writable(answer):
                root = os.path.abspath(answer)
                CONFIG["output_root"] = root
                save_config(CONFIG)
                print(f"已改為：{root}")
            else:
                print(f"[警告] 無法寫入 {answer}，仍使用 {root}")

    RAW_DIR = os.path.join(root, "初始檔案")
    PROCESSED_DIR = os.path.join(root, "修改過的")


# ---------- 網頁選擇器 ----------
SELECTORS = {
    "type_select": [
        "select:has(option:has-text('tiltDot'))",
        "select#select1",
        "select:below(:text('Instrument Type'))",
        "select:near(:text('儀器種類'))",
        "select",
    ],
    "sn_input": [
        "input[placeholder='序號']",
        "input[name*='serial' i]",
        "input:below(:text('serial number'))",
        "input",
    ],
    "download_button": [
        "button:has-text('download')",
        "a:has-text('download')",
        "button:has-text('下載')",
        "a:has-text('下載')",
    ],
}


# ---------- 輸入解析 ----------
def parse_serials(inp: str):
    """把「36901-36905, 36910」解析成序號清單（保留順序、去重）。"""
    out = []
    for part in (inp or "").split(","):
        part = part.strip()
        if not part:
            continue
        m = re.fullmatch(r"(\d+)\s*-\s*(\d+)", part)
        if m:
            start, end = map(int, m.groups())
            if start > end:
                print(f"[提醒] 範圍 {part} 起訖顛倒，已自動對調。")
                start, end = end, start
            out += [str(i) for i in range(start, end + 1)]
        else:
            out.append(part)

    seen = set()
    uniq = []
    for sn in out:
        if sn not in seen:
            seen.add(sn)
            uniq.append(sn)
    if len(uniq) != len(out):
        print("[提醒] 序號有重複，已自動去除。")
    return uniq


def normalize_date(s):
    """把 2026/8/12、2026-8-12 統一成 2026-08-12；無法解析則原樣回傳。"""
    s = (s or "").strip().replace("/", "-").replace(".", "-")
    m = re.fullmatch(r"(\d{4})-(\d{1,2})-(\d{1,2})", s)
    if m:
        y, mo, d = map(int, m.groups())
        try:
            return datetime(y, mo, d).strftime("%Y-%m-%d")
        except ValueError:
            return s
    return s


def normalize_time(s):
    """把 7:45 / 07:45 統一成 07:45:00；無法解析則原樣回傳。"""
    s = (s or "").strip()
    m = re.fullmatch(r"(\d{1,2}):(\d{1,2})(?::(\d{1,2}))?", s)
    if m:
        h, mi = int(m.group(1)), int(m.group(2))
        se = int(m.group(3) or 0)
        if h < 24 and mi < 60 and se < 60:
            return f"{h:02d}:{mi:02d}:{se:02d}"
    return s


def _valid_date(s) -> bool:
    try:
        datetime.strptime(s, "%Y-%m-%d")
        return True
    except (TypeError, ValueError):
        return False


def _valid_time(s) -> bool:
    try:
        datetime.strptime(s, "%H:%M:%S")
        return True
    except (TypeError, ValueError):
        return False


def ask_date(prompt: str) -> str:
    while True:
        val = normalize_date(input(prompt))
        if _valid_date(val):
            return val
        print("日期格式不正確，請輸入 YYYY-MM-DD（例如 2026-08-12）。")


def ask_time(prompt: str) -> str:
    while True:
        val = normalize_time(input(prompt))
        if _valid_time(val):
            return val
        print("時間格式不正確，請輸入 HH:MM 或 HH:MM:SS（例如 07:45）。")


def parse_col_input(prompt: str) -> int:
    """讀取欄位編號，接受數字 (5) 或 Excel 字母 (E)，回傳 0-based 索引。"""
    while True:
        val = input(prompt).strip()
        if val.isdigit():
            n = int(val)
            if n >= 1:
                return n - 1
            print("欄位編號要從 1 開始，請重新輸入。")
        elif val.isalpha() and val.isascii():
            try:
                return column_index_from_string(val.upper()) - 1
            except Exception:
                print("字母解析失敗，請輸入正確的 Excel 欄位字母 (例如 E)。")
        else:
            print("輸入格式錯誤，請輸入數字 (例如 5) 或英文字母 (例如 E)。")


# ---------- 網頁操作 ----------
async def click_first(page, sels, timeout=1500):
    for sel in sels:
        try:
            el = await page.wait_for_selector(sel, timeout=timeout)
            await el.click()
            return True
        except Exception:
            continue
    return False


async def select_tiltdot(page):
    for sel in SELECTORS["type_select"]:
        try:
            await page.wait_for_selector(sel, timeout=CONFIG["default_wait_ms"])
            await page.select_option(sel, label="tiltDot")
            return True
        except Exception:
            continue
    return False


async def fill_dates_once(page, start, end):
    async def set_val(sels, val):
        for sel in sels:
            try:
                loc = page.locator(sel).first
                await loc.wait_for(timeout=1500)
                await loc.evaluate("node => node.value = ''")
                await loc.fill(val)
                await loc.evaluate("node => node.dispatchEvent(new Event('change', { bubbles: true }))")
                return True
            except Exception:
                continue
        return False

    start_sels = [
        "input#date",
        "input:near(:text('Date (start)'))",
        "input[placeholder*='start' i]",
        "xpath=(//input[contains(@id, 'date') or contains(@class, 'date')])[1]",
    ]
    end_sels = [
        "input#date2",
        "input[name='date2']",
        "input:near(:text('Date (end)'))",
        "input:near(:text('迄'))",
        "input#edate",
        "input[placeholder*='end' i]",
        "xpath=(//input[contains(@id, 'date') or contains(@class, 'date')])[2]",
    ]

    ok1 = await set_val(start_sels, normalize_date(start))
    ok2 = await set_val(end_sels, normalize_date(end))
    return ok1 and ok2


async def trigger_query(page):
    return await page.evaluate("""() => {
        const btn = document.querySelector('button.submit');
        if (!btn) return {ok:false};
        btn.scrollIntoView();
        if (window.jQuery) window.jQuery(btn).trigger('click');
        else ['mousedown','mouseup','click'].forEach(e=>btn.dispatchEvent(new MouseEvent(e,{bubbles:true})));
        return {ok:true};
    }""")


# ---------- 原始檔解析 ----------
def _try_decode(data: bytes):
    for e in ("utf-8-sig", "utf-8", "cp950", "big5"):
        try:
            return data.decode(e), e
        except Exception:
            continue
    return data.decode("utf-8", "replace"), "utf-8"


def _detect_delim(t: str) -> str:
    """挑出最像分隔符號的字元。

    以「各列欄位數是否一致」評分，而不是單純數出現次數，
    避免資料用分號分隔、數值卻含小數逗號時誤判成逗號。
    """
    lines = [ln for ln in t.splitlines() if ln.strip()][:50]
    if not lines:
        return ","
    best, best_score = ",", -1.0
    for d in (",", ";", "\t", "|"):
        counts = [ln.count(d) for ln in lines]
        mode = max(set(counts), key=counts.count)
        if mode == 0:
            continue
        score = counts.count(mode) / len(counts) * mode
        if score > best_score:
            best_score, best = score, d
    return best


_DATE_RE = re.compile(r"^(\d{1,4})[/\-.](\d{1,2})[/\-.](\d{1,4})$")
_TIME_RE = re.compile(r"^(\d{1,2}):(\d{1,2})(?::(\d{1,2}))?")


def _cell(parts, col_1based):
    """取第 N 欄（1-based）並去掉引號與空白。"""
    idx = col_1based - 1
    if idx < 0 or idx >= len(parts):
        return ""
    return parts[idx].replace('"', "").strip()


def _parse_data_date(date_val: str, order: str):
    m = _DATE_RE.match(date_val)
    if not m:
        return None
    a, b, c = m.groups()
    if len(a) == 4:
        y, mo, d = int(a), int(b), int(c)
    elif order == "MDY":
        mo, d, y = int(a), int(b), int(c)
    else:
        d, mo, y = int(a), int(b), int(c)
    if y < 100:
        y += 2000
    try:
        return datetime(y, mo, d)
    except ValueError:
        return None


def _detect_date_order(lines, delim, date_col):
    """判斷原始檔的日期是 DD/MM/YYYY 還是 MM/DD/YYYY。"""
    saw_two_digit_first = False
    for line in lines:
        val = _cell(line.split(delim), date_col)
        m = _DATE_RE.match(val)
        if not m:
            continue
        a, b, _ = m.groups()
        if len(a) == 4:
            return "YMD"
        saw_two_digit_first = True
        if int(a) > 12:
            return "DMY"
        if int(b) > 12:
            return "MDY"
    if saw_two_digit_first:
        print("[提醒] 原始檔日期為 xx/xx/yyyy 且日、月皆 ≤ 12，無法自動判定順序，"
              "本次以「日/月/年」處理。若不正確，請把設定檔的 date_order 改成 \"MDY\"。")
    return "DMY"


def _line_datetime(line, delim, date_col, time_col, order):
    parts = line.split(delim)
    date_val = _cell(parts, date_col)
    time_val = _cell(parts, time_col)
    if not date_val or ":" not in time_val:
        return None
    d = _parse_data_date(date_val, order)
    if d is None:
        return None
    tm = _TIME_RE.match(time_val)
    if not tm:
        return None
    h, mi = int(tm.group(1)), int(tm.group(2))
    se = int(tm.group(3) or 0)
    if h > 23 or mi > 59 or se > 59:
        return None
    return d.replace(hour=h, minute=mi, second=se)


def _filter_data_lines(in_path, start_dt: datetime, end_dt: datetime):
    """只保留時間戳落在 [start_dt, end_dt] 之內的資料列。

    回傳 (文字, 分隔符號, 統計資訊)。時間比較一律用 datetime 物件，
    不再用字串大小比較（未補零的 2026/8/12 會讓字串比較整段失效）。
    """
    with open(in_path, "rb") as f:
        raw = f.read()

    text, _ = _try_decode(raw)
    text = text.replace("\x00", "").lstrip("﻿").lstrip("ï»¿")
    lines = text.splitlines()
    delim = _detect_delim(text)

    date_col = CONFIG["date_col"]
    time_col = CONFIG["time_col"]
    order = CONFIG.get("date_order", "auto")
    if str(order).upper() == "AUTO":
        order = _detect_date_order(lines, delim, date_col)
    else:
        order = str(order).upper()

    kept, total = [], 0
    for line in lines:
        dt = _line_datetime(line, delim, date_col, time_col, order)
        if dt is None:
            continue  # 表頭或格式不符的列
        total += 1
        if start_dt <= dt <= end_dt:
            kept.append(line)

    info = {"total": total, "kept": len(kept), "order": order}
    return "\n".join(kept), delim, info


# ---------- 換算與繪圖 ----------
def _excel_region_pixel_size(ws, first_col, last_col, first_row, last_row):
    total_w = 0
    total_h = 0
    for c in range(first_col, last_col + 1):
        w = ws.column_dimensions[get_column_letter(c)].width or 8.43
        total_w += w * 7 + 5
    for r in range(first_row, last_row + 1):
        h = ws.row_dimensions[r].height or 15
        total_h += h * 96 / 72
    return int(total_w), int(total_h)


def _numeric(df, col):
    return pd.to_numeric(df.iloc[:, col], errors="coerce")


def _first_valid_index(df, cols):
    """第一列「所有指定欄位都是數值」的索引；圖表與 Excel 公式共用同一個基準。"""
    if df.empty:
        return None
    mask = None
    for c in cols:
        if c is None:
            continue
        s = pd.notna(_numeric(df, c))
        mask = s if mask is None else (mask & s)
    if mask is None or not mask.any():
        return None
    return mask.idxmax()


def _safe_coef(c, name):
    """補償係數若為 nan/inf，寫進 Excel 公式會讓檔案開不起來。"""
    try:
        c = float(c)
    except (TypeError, ValueError):
        c = float("nan")
    if not np.isfinite(c):
        print(f"[警告] {name} 補償係數無法計算（非有限值），已視為 0。")
        return 0.0
    return c


def _compute_axes(df, col_x, col_y, col_z, axes_type, base_idx):
    E = _numeric(df, col_x)
    F = _numeric(df, col_y)

    base_E = float(E.loc[base_idx]) if base_idx is not None else 0.0
    base_F = float(F.loc[base_idx]) if base_idx is not None else 0.0

    X_val = (E - base_E) * 3600
    Y_val = (F - base_F) * 3600

    if axes_type == "XYZ":
        G = _numeric(df, col_z)
        base_G = float(G.loc[base_idx]) if base_idx is not None else 0.0
        return X_val, Y_val, (G - base_G) * 3600
    return X_val, Y_val, None


def _row_stamp(df, idx):
    """組出異常紀錄用的「日期 時間」字串。"""
    d_i = CONFIG["date_col"] - 1
    t_i = CONFIG["time_col"] - 1
    parts = []
    for i in (d_i, t_i):
        if 0 <= i < df.shape[1]:
            parts.append(str(df.iloc[idx, i]).strip().replace('"', ""))
    return " ".join(p for p in parts if p)


def _plot_to_png(df, label, right_center, axes_type, col_x, col_y, col_z, col_t, base_idx):
    """未補償圖表。回傳 (圖片, 異常紀錄, 異常列索引集合)。"""
    if df.empty:
        return None, [], set()

    vals = _compute_axes(df, col_x, col_y, col_z, axes_type, base_idx)
    if axes_type == "XYZ":
        X_val, Y_val, Z_val = vals
        s_vals = [X_val, Y_val, Z_val]
        names = ["X", "Y", "Z"]
    else:
        X_val, Y_val, _ = vals
        Z_val = None
        s_vals = [X_val, Y_val]
        names = ["X", "Y"]

    # --- 異常值偵測與過濾 ---
    outliers_record = []
    bad_indices = set()
    if CONFIG["remove_outliers"]:
        threshold = CONFIG["outlier_threshold"]
        for axis_name, s_val in zip(names, s_vals):
            bad_idx = s_val.index[s_val.abs() > threshold].tolist()
            for idx in bad_idx:
                bad_indices.add(idx)
                val = s_val.loc[idx]
                stamp = _row_stamp(df, idx)
                pos = f"Excel 第 {idx + 2} 列"
                where = f"於 {stamp} ({pos})" if stamp else f"於 {pos}"
                outliers_record.append(f"{axis_name}軸 {where} 發現異常數值: {val:.1f}")
                s_val.loc[idx] = np.nan

    finite = pd.concat(s_vals).replace([np.inf, -np.inf], np.nan).dropna()

    T_val = _numeric(df, col_t)
    X = range(len(df))

    y_min, y_max = (-400, 400) if finite.empty else (finite.min(), finite.max())
    pad = max(0.05 * (y_max - y_min), 40)
    y_min -= pad
    y_max += pad
    step = 40
    y_start = math.floor(y_min / step) * step
    y_end = math.ceil(y_max / step) * step

    right_center = _resolve_center(right_center, T_val)

    fig, ax1 = plt.subplots()
    try:
        ax2 = ax1.twinx()

        ax1.plot(X, X_val, label="X (sec)")
        ax1.plot(X, Y_val, label="Y (sec)")
        if axes_type == "XYZ":
            ax1.plot(X, Z_val, label="Z (sec)")

        for y in np.arange(y_start, y_end + step, step):
            ax1.axhline(y, color="gray" if abs(y) < 1e-6 else "#ddd",
                        linestyle="--" if abs(y) < 1e-6 else ":")
        ax1.set_ylim(y_start, y_end)

        rmin, rmax = _right_axis_range(y_start, y_end, T_val, right_center)
        ax2.set_ylim(rmin, rmax)
        ax2.plot(X, T_val, color="red", linestyle="--", label="T (°C)")

        ax1.set_xlabel(label)
        ax1.set_ylabel("X/Y/Z (sec)" if axes_type == "XYZ" else "X/Y (sec)")
        ax2.set_ylabel("T (°C)")

        lines_1, labels_1 = ax1.get_legend_handles_labels()
        lines_2, labels_2 = ax2.get_legend_handles_labels()
        ax1.legend(lines_1 + lines_2, labels_1 + labels_2, loc="upper right")

        peak_texts = []
        arr_t = T_val.to_numpy(dtype=float)
        if not np.isnan(arr_t).all():
            v_max_t = np.nanmax(arr_t)
            v_min_t = np.nanmin(arr_t)
            peak_t = v_max_t if abs(v_max_t) >= abs(v_min_t) else v_min_t
            peak_texts.append(f"T Peak: {peak_t:.1f} °C")

        for name, s_val in zip(names, s_vals):
            arr = s_val.to_numpy(dtype=float)
            if np.isnan(arr).all():
                continue
            v_max = np.nanmax(arr)
            v_min = np.nanmin(arr)
            peak = v_max if abs(v_max) >= abs(v_min) else v_min
            peak_idx = int(np.nanargmax(np.where(np.isnan(arr), -np.inf, np.abs(arr))))

            peak_t_val = T_val.iloc[peak_idx]
            t_diff = peak_t_val - right_center if np.isfinite(peak_t_val) else 0.0
            error_val = peak / t_diff if abs(t_diff) > 1e-6 else 0.0
            peak_texts.append(f"{name} Peak: {peak:.1f} | ΔT: {t_diff:.1f} | Error: {error_val:.4f}")

        _draw_peak_box(ax1, peak_texts)

        bio = BytesIO()
        fig.savefig(bio, dpi=150, format="png")
        bio.seek(0)
    finally:
        plt.close(fig)
    return bio, outliers_record, bad_indices


def _plot_comp_to_png(df, comp_X, comp_Y, comp_Z, T_raw, label, right_center, axes_type, bad_indices):
    """補償後圖表。"""
    X = range(len(df))

    comp_X_plot = comp_X.copy()
    comp_Y_plot = comp_Y.copy()
    for idx in bad_indices:
        if idx in comp_X_plot.index:
            comp_X_plot.loc[idx] = np.nan
            comp_Y_plot.loc[idx] = np.nan

    if axes_type == "XYZ":
        comp_Z_plot = comp_Z.copy()
        for idx in bad_indices:
            if idx in comp_Z_plot.index:
                comp_Z_plot.loc[idx] = np.nan
        s_vals = [comp_X_plot, comp_Y_plot, comp_Z_plot]
        names = ["Comp X", "Comp Y", "Comp Z"]
    else:
        comp_Z_plot = None
        s_vals = [comp_X_plot, comp_Y_plot]
        names = ["Comp X", "Comp Y"]

    finite = pd.concat(s_vals).replace([np.inf, -np.inf], np.nan).dropna()

    y_min, y_max = (-40, 40) if finite.empty else (finite.min(), finite.max())
    pad = max(0.05 * (y_max - y_min), 10)
    y_min -= pad
    y_max += pad

    y_range = y_max - y_min
    if y_range > 100:
        step = 40
    elif y_range > 50:
        step = 20
    elif y_range > 20:
        step = 10
    else:
        step = 5

    y_start = math.floor(y_min / step) * step
    y_end = math.ceil(y_max / step) * step

    right_center = _resolve_center(right_center, T_raw)

    fig, ax1 = plt.subplots()
    try:
        ax2 = ax1.twinx()

        ax1.plot(X, comp_X_plot, label="Comp X (sec)")
        ax1.plot(X, comp_Y_plot, label="Comp Y (sec)")
        if axes_type == "XYZ":
            ax1.plot(X, comp_Z_plot, label="Comp Z (sec)")

        for y in np.arange(y_start, y_end + step, step):
            ax1.axhline(y, color="gray" if abs(y) < 1e-6 else "#ddd",
                        linestyle="--" if abs(y) < 1e-6 else ":")
        ax1.set_ylim(y_start, y_end)

        rmin, rmax = _right_axis_range(y_start, y_end, T_raw, right_center)
        ax2.set_ylim(rmin, rmax)
        ax2.plot(X, T_raw, color="red", linestyle="--", label="T (°C)")

        ax1.set_xlabel(f"{label} (Compensated)")
        ax1.set_ylabel("Comp X/Y/Z (sec)" if axes_type == "XYZ" else "Comp X/Y (sec)")
        ax2.set_ylabel("T (°C)")

        lines_1, labels_1 = ax1.get_legend_handles_labels()
        lines_2, labels_2 = ax2.get_legend_handles_labels()
        ax1.legend(lines_1 + lines_2, labels_1 + labels_2, loc="upper right")

        peak_texts = []
        arr_t = T_raw.to_numpy(dtype=float)
        if not np.isnan(arr_t).all():
            v_max_t = np.nanmax(arr_t)
            v_min_t = np.nanmin(arr_t)
            peak_t = v_max_t if abs(v_max_t) >= abs(v_min_t) else v_min_t
            peak_texts.append(f"T Peak: {peak_t:.1f} °C")

        for name, s_val in zip(names, s_vals):
            arr = s_val.to_numpy(dtype=float)
            if np.isnan(arr).all():
                continue
            v_max = np.nanmax(arr)
            v_min = np.nanmin(arr)
            peak = v_max if abs(v_max) >= abs(v_min) else v_min
            peak_texts.append(f"{name} Peak: {peak:.1f}")

        _draw_peak_box(ax1, peak_texts)

        bio = BytesIO()
        fig.savefig(bio, dpi=150, format="png")
        bio.seek(0)
    finally:
        plt.close(fig)
    return bio


def _resolve_center(right_center, T_series):
    if right_center is not None and np.isfinite(right_center):
        return float(right_center)
    arr = T_series.to_numpy(dtype=float)
    if len(arr) and np.isfinite(arr[0]):
        return float(arr[0])
    if len(arr) and not np.isnan(arr).all():
        return float(np.nanmean(arr))
    return 0.0


def _right_axis_range(y_start, y_end, T_series, right_center):
    """讓右軸的溫度基準線對齊左軸的 0 線。"""
    arr = T_series.to_numpy(dtype=float)
    frac = (0 - y_start) / max(y_end - y_start, 1e-9)
    frac = min(max(frac, 0.05), 0.95)
    if len(arr) == 0 or np.isnan(arr).all():
        du = dv = 1.0
    else:
        du = float(np.nanmax(arr - right_center))
        dv = float(np.nanmax(right_center - arr))
    du = max(du, 1.0)
    dv = max(dv, 1.0)
    ratio = (1 - frac) / frac
    down = max(dv, du / max(ratio, 1e-9))
    up = max(du, ratio * down)
    return right_center - down * 1.05, right_center + up * 1.05


def _draw_peak_box(ax, peak_texts):
    if not peak_texts:
        return
    props = dict(boxstyle="round,pad=0.5", facecolor="white", alpha=0.8, edgecolor="gray")
    ax.text(0.97, 0.05, "\n".join(peak_texts), transform=ax.transAxes, fontsize=10,
            verticalalignment="bottom", horizontalalignment="right",
            multialignment="left", bbox=props, zorder=10)


def _write_cell(ws, row, col, value):
    """寫入儲存格；以 = 開頭的文字要強制當成文字，否則 Excel 會當公式而報錯。"""
    cell = ws.cell(row, col, value=value)
    if isinstance(value, str) and value.startswith("="):
        cell.data_type = "s"
    return cell


def transform_text_to_xlsx(filtered_text, delim, in_path, axes_type, col_x, col_y, col_z, col_t,
                           run_mode="TRANSFER"):
    lines = filtered_text.splitlines()
    if not lines:
        return None, [], None

    max_cols = max(len(line.split(delim)) for line in lines)
    df = pd.read_csv(StringIO(filtered_text), sep=delim, engine="python",
                     header=None, names=range(max_cols))
    df = df.fillna("")

    basename = os.path.basename(in_path)
    m = re.search(r"(\d{3,})", basename)
    serial = m.group(1) if m else "Index"

    target_cols = [col_x, col_y, col_t] + ([col_z] if axes_type == "XYZ" else [])
    for c in target_cols:
        if c >= df.shape[1]:
            raise ValueError(f"欄位設定超出原始資料範圍：原始檔只有 {df.shape[1]} 欄，"
                             f"卻指定了第 {c + 1} 欄。請重新確認 X/Y/Z/T 的欄位。")

    base_idx = _first_valid_index(df, target_cols)
    if base_idx is None:
        raise ValueError("指定的 X/Y/Z/T 欄位在範圍內找不到任何數值資料，請確認欄位設定是否正確。")

    T_series = _numeric(df, col_t)
    t_center = float(T_series.loc[base_idx])

    # --- 1. 未補償圖表，同時取得異常列 ---
    img_bytes, outliers_record, bad_indices = None, [], set()
    try:
        img_bytes, outliers_record, bad_indices = _plot_to_png(
            df, label=serial, right_center=t_center, axes_type=axes_type,
            col_x=col_x, col_y=col_y, col_z=col_z, col_t=col_t, base_idx=base_idx)
    except Exception as e:
        print(f"[PLOT] 嵌圖失敗：{e}")

    # --- 補償係數（僅 TEST 模式）---
    c_X = c_Y = c_Z = 0.0
    img_bytes_comp = None
    comp_mode = (run_mode == "TEST")

    if comp_mode:
        try:
            tmp_T = _numeric(df, col_t)
            tmp_X = _numeric(df, col_x)
            tmp_Y = _numeric(df, col_y)
            tmp_Z = _numeric(df, col_z) if axes_type == "XYZ" else None

            valid_mask = ~df.index.isin(bad_indices) & pd.notna(tmp_T) & pd.notna(tmp_X) & pd.notna(tmp_Y)
            if axes_type == "XYZ":
                valid_mask &= pd.notna(tmp_Z)

            T_base_val = float(tmp_T.loc[base_idx])
            X_base_val = float(tmp_X.loc[base_idx])
            Y_base_val = float(tmp_Y.loc[base_idx])
            Z_base_val = float(tmp_Z.loc[base_idx]) if axes_type == "XYZ" else 0.0

            c_X = _safe_coef(_c_from_peak(tmp_X, X_base_val, tmp_T, T_base_val, valid_mask), "X軸")
            c_Y = _safe_coef(_c_from_peak(tmp_Y, Y_base_val, tmp_T, T_base_val, valid_mask), "Y軸")
            if axes_type == "XYZ":
                c_Z = _safe_coef(_c_from_peak(tmp_Z, Z_base_val, tmp_T, T_base_val, valid_mask), "Z軸")

            comp_X_s = (tmp_X - X_base_val + c_X * (tmp_T - T_base_val)) * 3600
            comp_Y_s = (tmp_Y - Y_base_val + c_Y * (tmp_T - T_base_val)) * 3600
            comp_Z_s = (tmp_Z - Z_base_val + c_Z * (tmp_T - T_base_val)) * 3600 if axes_type == "XYZ" else None
            img_bytes_comp = _plot_comp_to_png(df, comp_X_s, comp_Y_s, comp_Z_s, tmp_T,
                                               label=serial, right_center=t_center,
                                               axes_type=axes_type, bad_indices=bad_indices)
        except Exception as e:
            print(f"[PLOT COMP] 補償計算或圖表產生失敗：{e}")
            c_X = c_Y = c_Z = 0.0

    # --- 2. 寫入 Excel ---
    df2 = pd.concat([pd.DataFrame([[""] * df.shape[1]], columns=df.columns), df], ignore_index=True)

    offset = 2
    out_col_x = max_cols + offset
    out_col_y = out_col_x + 1
    if axes_type == "XYZ":
        out_col_z = out_col_y + 1
        out_col_t_diff = out_col_z + 1
    else:
        out_col_z = None
        out_col_t_diff = out_col_y + 1

    if comp_mode:
        out_col_cx = out_col_t_diff + 1
        out_col_cy = out_col_cx + 1
        out_col_cz = out_col_cy + 1 if axes_type == "XYZ" else None
        needed_cols = (out_col_cz if axes_type == "XYZ" else out_col_cy) + 1
    else:
        out_col_cx = out_col_cy = out_col_cz = None
        needed_cols = out_col_t_diff + 1

    if df2.shape[1] < needed_cols:
        for i in range(df2.shape[1], needed_cols):
            df2[i] = ""

    df2.iloc[0, col_x] = "X"
    df2.iloc[0, col_y] = "Y"
    if axes_type == "XYZ":
        df2.iloc[0, col_z] = "Z"
    df2.iloc[0, col_t] = "T"

    df2.iloc[0, out_col_x] = "X(sec)"
    df2.iloc[0, out_col_y] = "Y(sec)"
    if axes_type == "XYZ":
        df2.iloc[0, out_col_z] = "Z(sec)"
    df2.iloc[0, out_col_t_diff] = "ΔT"

    if comp_mode:
        df2.iloc[0, out_col_cx] = "補償後X(sec)"
        df2.iloc[0, out_col_cy] = "補償後Y(sec)"
        if axes_type == "XYZ":
            df2.iloc[0, out_col_cz] = "補償後Z(sec)"

    wb = Workbook()
    ws = wb.active
    ws.title = "data"

    for r in range(df2.shape[0]):
        df_idx = r - 1
        is_bad = df_idx in bad_indices
        for c in range(df2.shape[1]):
            value = "" if (is_bad and c in target_cols) else df2.iat[r, c]
            cell = _write_cell(ws, r + 1, c + 1, value)
            if r == 0:
                cell.alignment = Alignment(horizontal="right")

    letter_x = get_column_letter(col_x + 1)
    letter_y = get_column_letter(col_y + 1)
    letter_z = get_column_letter(col_z + 1) if axes_type == "XYZ" else None
    letter_t = get_column_letter(col_t + 1)

    out_letter_x = get_column_letter(out_col_x + 1)
    out_letter_y = get_column_letter(out_col_y + 1)
    out_letter_z = get_column_letter(out_col_z + 1) if axes_type == "XYZ" else None
    out_letter_t_diff = get_column_letter(out_col_t_diff + 1)
    out_letter_cx = get_column_letter(out_col_cx + 1) if comp_mode else None
    out_letter_cy = get_column_letter(out_col_cy + 1) if comp_mode else None
    out_letter_cz = get_column_letter(out_col_cz + 1) if (comp_mode and axes_type == "XYZ") else None

    # 圖表與公式共用同一個基準列
    base_row = int(base_idx) + 2
    base_x = f"${letter_x}${base_row}"
    base_y = f"${letter_y}${base_row}"
    base_z = f"${letter_z}${base_row}" if axes_type == "XYZ" else None
    base_t = f"${letter_t}${base_row}"

    for i in range(2, df2.shape[0] + 1):
        df_idx = i - 2
        if df_idx in bad_indices:
            cells = [out_letter_x, out_letter_y, out_letter_t_diff]
            if axes_type == "XYZ":
                cells.append(out_letter_z)
            if comp_mode:
                cells += [out_letter_cx, out_letter_cy]
                if axes_type == "XYZ":
                    cells.append(out_letter_cz)
            for letter in cells:
                ws[f"{letter}{i}"] = ""
            continue

        ws[f"{out_letter_t_diff}{i}"] = f"={letter_t}{i}-{base_t}"
        ws[f"{out_letter_x}{i}"] = f"=({letter_x}{i}-{base_x})*3600"
        ws[f"{out_letter_y}{i}"] = f"=({letter_y}{i}-{base_y})*3600"
        if axes_type == "XYZ":
            ws[f"{out_letter_z}{i}"] = f"=({letter_z}{i}-{base_z})*3600"

        if comp_mode:
            ws[f"{out_letter_cx}{i}"] = f"=({letter_x}{i}-{base_x}+({c_X:.8f}*{out_letter_t_diff}{i}))*3600"
            ws[f"{out_letter_cy}{i}"] = f"=({letter_y}{i}-{base_y}+({c_Y:.8f}*{out_letter_t_diff}{i}))*3600"
            if axes_type == "XYZ":
                ws[f"{out_letter_cz}{i}"] = f"=({letter_z}{i}-{base_z}+({c_Z:.8f}*{out_letter_t_diff}{i}))*3600"

    # 圖表放在資料右側的空白區，不覆蓋資料
    w, h = _excel_region_pixel_size(ws, 1, needed_cols, 3, 30)
    anchor_col = get_column_letter(needed_cols + 2)
    if img_bytes:
        xl_img = XLImage(PILImage.open(img_bytes))
        xl_img.width = int(w * 0.5)
        xl_img.height = int(h * 1.1)
        ws.add_image(xl_img, f"{anchor_col}3")

    if img_bytes_comp:
        xl_img_comp = XLImage(PILImage.open(img_bytes_comp))
        xl_img_comp.width = int(w * 0.5)
        xl_img_comp.height = int(h * 1.1)
        ws.add_image(xl_img_comp, f"{anchor_col}35")

    os.makedirs(PROCESSED_DIR, exist_ok=True)
    out_filename = os.path.splitext(basename)[0] + ".xlsx"
    if outliers_record:
        out_filename = "[異常]_" + out_filename

    out = os.path.join(PROCESSED_DIR, out_filename)
    wb.save(out)
    return out, outliers_record, base_idx


def _c_from_peak(s_raw, base_raw, t_raw, t_base, valid_mask):
    """以「最大峰值點與基準點連線」求該軸的溫度補償係數。"""
    s_sec_valid = ((s_raw - base_raw) * 3600)[valid_mask]
    if s_sec_valid.empty:
        return 0.0
    v_max = s_sec_valid.max()
    v_min = s_sec_valid.min()
    peak = v_max if abs(v_max) >= abs(v_min) else v_min
    peak_idx = s_sec_valid[s_sec_valid == peak].index[0]
    t_diff = t_raw.loc[peak_idx] - t_base
    if abs(t_diff) > 1e-6:
        return -((s_raw.loc[peak_idx] - base_raw) / t_diff)
    return 0.0


# ---------- 單序號流程 ----------
async def process_one_sn(page, sn, start_dt, end_dt, axes_type, col_x, col_y, col_z, col_t, run_mode):
    """回傳 dict：status = ok / no_data / convert_failed。"""
    result = {"sn": sn, "status": "no_data", "outliers": [], "xlsx": None, "reason": ""}

    filled = False
    for sel in SELECTORS["sn_input"]:
        try:
            loc = page.locator(sel).first
            await loc.wait_for(timeout=CONFIG["default_wait_ms"])
            await loc.evaluate("node => node.value = ''")
            await loc.fill(sn)
            await page.wait_for_timeout(300)
            filled = True
            break
        except Exception:
            continue

    if not filled:
        # 不能繼續查詢，否則網頁還停在上一個序號，會下載到別人的資料
        print(f"[{sn}] 序號欄位填入失敗，略過（避免下載到上一筆序號的資料）")
        result["reason"] = "序號欄位填入失敗"
        return result

    await trigger_query(page)
    try:
        async with page.expect_download(timeout=CONFIG["download_timeout_ms"]) as dl:
            await click_first(page, SELECTORS["download_button"])
        d = await dl.value
    except Exception:
        print(f"[{sn}] 無資料或逾時")
        result["reason"] = "下載逾時或網站無回應"
        return result

    os.makedirs(RAW_DIR, exist_ok=True)
    filename = d.suggested_filename or "download.txt"
    path = os.path.join(RAW_DIR, f"{sn}_{filename}")
    await d.save_as(path)

    if path.lower().endswith(".txt") and os.path.getsize(path) < 100:
        print(f"[{sn}] 無資料，略過空白檔案：{os.path.basename(path)}")
        result["reason"] = "下載檔案為空"
        return result

    print(f"[{sn}] 已下載原始檔案至：{path}")

    try:
        filtered_text, delim, info = _filter_data_lines(path, start_dt, end_dt)
        print(f"[{sn}] 原始資料 {info['total']} 筆，時間範圍內 {info['kept']} 筆")

        if not filtered_text.strip():
            print(f"[{sn}] 指定的日期時間範圍內無有效資料")
            result["reason"] = "時間範圍內無資料"
            return result

        xlsx, outliers, _ = transform_text_to_xlsx(
            filtered_text, delim, path, axes_type, col_x, col_y, col_z, col_t, run_mode)
        if xlsx:
            print(f"[{sn}] 已產生 Excel 檔案至：{xlsx}")
            result.update(status="ok", outliers=outliers, xlsx=xlsx)
            return result
        result["status"] = "convert_failed"
        result["reason"] = "轉換後無資料"
    except Exception as e:
        print(f"[{sn}] 轉換失敗：{e}")
        traceback.print_exc()
        result["status"] = "convert_failed"
        result["reason"] = str(e)

    return result


# ---------- 校正參數總表 ----------
def generate_calibration_parameters(file_list, out_dir, axes_type, col_x, col_y, col_z, col_t):
    """只統計本次產出的檔案，避免混入上一次執行留下的舊 xlsx。"""
    all_sensor_params = []

    print("\n==========================================")
    print("開始萃取各感測器回歸參數 (基於最大傾斜峰值)...\n")

    for file_path in file_list:
        filename = os.path.basename(file_path)
        if filename.startswith("~$") or not os.path.exists(file_path):
            continue

        print(f"正在分析檔案: {filename}")
        try:
            df = pd.read_excel(file_path, sheet_name=0)

            cols = [col_x, col_y, col_t] + ([col_z] if axes_type == "XYZ" else [])
            if df.shape[1] <= max(cols):
                print("  [略過] 格式不符（欄位不足）。")
                continue

            T_raw = _numeric(df, col_t)
            X_raw = _numeric(df, col_x)
            Y_raw = _numeric(df, col_y)
            Z_raw = _numeric(df, col_z) if axes_type == "XYZ" else None

            valid_mask = pd.notna(T_raw) & pd.notna(X_raw) & pd.notna(Y_raw)
            if axes_type == "XYZ":
                valid_mask &= pd.notna(Z_raw)

            if not valid_mask.any():
                print("  [略過] 無有效數據。")
                continue

            base_idx = valid_mask.idxmax()
            T_base = float(T_raw.loc[base_idx])
            X_base = float(X_raw.loc[base_idx])
            Y_base = float(Y_raw.loc[base_idx])

            sensor_data = {
                "感測器檔名": filename,
                "基準溫度(T_base)": T_base,
                "X軸補償係數 (cx)": _safe_coef(_c_from_peak(X_raw, X_base, T_raw, T_base, valid_mask), "X軸"),
                "Y軸補償係數 (cy)": _safe_coef(_c_from_peak(Y_raw, Y_base, T_raw, T_base, valid_mask), "Y軸"),
            }
            if axes_type == "XYZ":
                Z_base = float(Z_raw.loc[base_idx])
                sensor_data["Z軸補償係數 (cz)"] = _safe_coef(
                    _c_from_peak(Z_raw, Z_base, T_raw, T_base, valid_mask), "Z軸")

            all_sensor_params.append(sensor_data)
            print("  參數計算成功 (基於最大傾斜峰值)")

        except Exception as e:
            print(f"  [錯誤] 處理時發生錯誤: {e}")

    if not all_sensor_params:
        print("\n沒有成功解析出任何參數。")
        return None

    results_df = pd.DataFrame(all_sensor_params)
    output_path = os.path.join(out_dir, "Linear_Calibration_Parameters.xlsx")
    try:
        results_df.to_excel(output_path, index=False)
    except PermissionError:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(out_dir, f"Linear_Calibration_Parameters_{stamp}.xlsx")
        print("[提醒] 原總表檔案可能正被 Excel 開啟，改存新檔。")
        try:
            results_df.to_excel(output_path, index=False)
        except Exception as e:
            print(f"[錯誤] 總表寫入失敗：{e}")
            return None
    except Exception as e:
        print(f"[錯誤] 總表寫入失敗：{e}")
        return None

    print(f"\n參數萃取完畢！總表：{os.path.basename(output_path)}")
    return output_path


# ---------- 報告 ----------
def write_report(all_outliers, no_data, convert_failed):
    os.makedirs(PROCESSED_DIR, exist_ok=True)
    report_path = os.path.join(PROCESSED_DIR, "異常資料過濾報告.txt")
    with open(report_path, "w", encoding="utf-8-sig") as f:
        f.write("==========================================\n")
        f.write("異常資料與無資料序號報告\n")
        f.write(f"產生時間：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("==========================================\n\n")

        if no_data:
            f.write("【範圍內無資料或下載失敗的序號】\n")
            for sn, reason in no_data:
                f.write(f"   - {sn}（{reason}）\n")
            f.write("\n")

        if convert_failed:
            f.write("【已下載但轉換失敗的序號】\n")
            for sn, reason in convert_failed:
                f.write(f"   - {sn}（{reason}）\n")
            f.write("\n")

        if all_outliers:
            f.write(f"【異常資料過濾紀錄 (超過閾值 ±{CONFIG['outlier_threshold']:.0f})】\n")
            for sn, recs in all_outliers.items():
                f.write(f"► 異常檔案序號: {sn} (共 {len(recs)} 筆)\n")
                for r in recs:
                    f.write(f"   - {r}\n")
                f.write("\n")
        f.write("==========================================\n")
    return report_path


# ---------- 主程式 ----------
def select_mode():
    print("=" * 42)
    print("請選擇執行模式：")
    print("  1. 資料轉換模式 (transfer) [預設]")
    print("     下載 → 換算 → 異常過濾 → 產出未補償圖表 → 產出校正係數總表")
    print("  2. 補償驗證模式 (test)")
    print("     在模式 1 之上，額外產出補償後欄位與補償後圖表")
    print("=" * 42)
    while True:
        val = input("輸入 1 或 2 (直接按 Enter 預設為 1)：").strip()
        if val in ("1", ""):
            return "TRANSFER"
        if val == "2":
            return "TEST"
        print("輸入錯誤，請輸入 1 或 2 (或直接按 Enter)。")


async def run():
    load_config()
    setup_output_dirs(interactive=True)

    run_mode = select_mode()
    print(f"目前模式：{'資料轉換 (transfer)' if run_mode == 'TRANSFER' else '補償驗證 (test)'}\n")

    sn_input = input("請輸入序號(例如 36901-36919)：")

    axes_type = input("請問包含的軸是 XYZ 還是 XY? (輸入 XYZ 或 XY，預設包含T): ").strip().upper()
    if axes_type not in ("XYZ", "XY"):
        axes_type = "XYZ"

    print("\n請輸入下載下來的原始資料中，各數值對應的行數 (可輸入數字如 5 或字母如 E):")
    col_x = parse_col_input("X軸 所在的行數 (例: 5 或 E): ")
    col_y = parse_col_input("Y軸 所在的行數 (例: 6 或 F): ")
    col_z = parse_col_input("Z軸 所在的行數 (例: 7 或 G): ") if axes_type == "XYZ" else None
    col_t = parse_col_input("T(溫度) 所在的行數 (例: 12 或 L): ")
    print("-" * 30)

    start_date = ask_date("資料起日期(例如 2026-08-12)：")
    start_time = ask_time("資料起時間(例如 07:45 或 07:45:00)：")
    end_date = ask_date("資料迄日期(例如 2026-08-13)：")
    end_time = ask_time("資料迄時間(例如 17:30 或 17:30:00)：")

    start_dt = datetime.strptime(f"{start_date} {start_time}", "%Y-%m-%d %H:%M:%S")
    end_dt = datetime.strptime(f"{end_date} {end_time}", "%Y-%m-%d %H:%M:%S")
    if end_dt < start_dt:
        print("[提醒] 起訖時間顛倒，已自動對調。")
        start_dt, end_dt = end_dt, start_dt

    serials = parse_serials(sn_input)
    if not serials:
        print("沒有可處理的序號，結束。")
        return

    if async_playwright is None:
        print("[錯誤] 找不到 Playwright 元件，無法連線下載。請重新安裝本程式。")
        return

    all_outliers = {}
    no_data = []
    convert_failed = []
    produced = []

    async with async_playwright() as p:
        try:
            browser = await p.chromium.launch(headless=CONFIG["headless"])
        except Exception as e:
            print("[錯誤] 瀏覽器元件無法啟動，請重新安裝本程式。")
            print(f"       詳細訊息：{e}")
            return

        try:
            ctx = await browser.new_context(accept_downloads=True)
            page = await ctx.new_page()

            await page.goto(CONFIG["url"])
            if not await select_tiltdot(page):
                print("[警告] 找不到儀器種類選單，網站可能已改版或需要先登入。")
            if not await fill_dates_once(page, start_date, end_date):
                print("[警告] 日期欄位填寫失敗，查詢結果可能不是預期的範圍。")

            for sn in serials:
                print(f"--- 處理 {sn} ---")
                res = await process_one_sn(page, sn, start_dt, end_dt, axes_type,
                                           col_x, col_y, col_z, col_t, run_mode)
                if res["status"] == "ok":
                    if res["outliers"]:
                        all_outliers[res["sn"]] = res["outliers"]
                    if res["xlsx"]:
                        produced.append(res["xlsx"])
                elif res["status"] == "convert_failed":
                    convert_failed.append((res["sn"], res["reason"]))
                else:
                    no_data.append((res["sn"], res["reason"]))
        finally:
            await browser.close()

    print("\n==========================================")
    if all_outliers or no_data or convert_failed:
        report_path = write_report(all_outliers, no_data, convert_failed)
        print(f"發現異常資料或無資料序號，報告已匯出：{report_path}")
    else:
        print("沒有發現資料異常檔案，且所有序號皆下載成功")
    print("==========================================\n")

    if produced:
        generate_calibration_parameters(produced, PROCESSED_DIR, axes_type,
                                        col_x, col_y, col_z, col_t)
    else:
        print("本次沒有產出任何 Excel 檔案，略過校正參數總表。")


def main():
    _reconfigure_stdio()
    try:
        asyncio.run(run())
    except KeyboardInterrupt:
        print("\n已中止。")
    except Exception:
        traceback.print_exc()
    finally:
        try:
            input("\n執行結束，按 Enter 關閉視窗...")
        except EOFError:
            pass


if __name__ == "__main__":
    main()
